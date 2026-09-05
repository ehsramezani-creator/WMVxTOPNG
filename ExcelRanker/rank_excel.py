
from pathlib import Path
from copy import copy
from openpyxl import load_workbook


# ============================================================
# File paths
# ============================================================

BASE_DIR = Path(__file__).resolve().parent

INPUT_FILE = BASE_DIR / "input.xlsx"
OUTPUT_FILE = BASE_DIR / "ranked.xlsx"


# ============================================================
# Convert cell value to number
# ============================================================

def numeric_value(value):
    if value is None:
        return 0

    if isinstance(value, (int, float)):
        return value

    try:
        return float(value)
    except (ValueError, TypeError):
        return 0


# ============================================================
# Read complete row
# ============================================================

def read_row(sheet, row_number):
    row_data = []

    for column_number in range(1, 5):  # A, B, C, D

        cell = sheet.cell(
            row=row_number,
            column=column_number
        )

        row_data.append({
            "value": cell.value,
            "style": copy(cell._style),
            "number_format": cell.number_format,
            "font": copy(cell.font),
            "fill": copy(cell.fill),
            "border": copy(cell.border),
            "alignment": copy(cell.alignment),
            "protection": copy(cell.protection),
        })

    return {
        "data": row_data,
        "height": sheet.row_dimensions[row_number].height,
        "hidden": sheet.row_dimensions[row_number].hidden,
    }


# ============================================================
# Write complete row
# ============================================================

def write_row(sheet, row_number, row_data):

    sheet.row_dimensions[row_number].height = row_data["height"]
    sheet.row_dimensions[row_number].hidden = row_data["hidden"]

    for column_number, cell_data in enumerate(
        row_data["data"],
        start=1
    ):

        cell = sheet.cell(
            row=row_number,
            column=column_number
        )

        cell.value = cell_data["value"]

        cell._style = copy(cell_data["style"])
        cell.number_format = cell_data["number_format"]

        cell.font = copy(cell_data["font"])
        cell.fill = copy(cell_data["fill"])
        cell.border = copy(cell_data["border"])
        cell.alignment = copy(cell_data["alignment"])
        cell.protection = copy(cell_data["protection"])


# ============================================================
# Sort Excel file
# ============================================================

def sort_excel(input_file, output_file):

    # --------------------------------------------------------
    # Load workbook
    # --------------------------------------------------------

    workbook = load_workbook(input_file)
    sheet = workbook.active

    print("========================================")
    print("Excel Ranker")
    print("========================================")
    print(f"Input : {input_file}")
    print(f"Output: {output_file}")
    print(f"Sheet : {sheet.title}")
    print(f"Rows  : {sheet.max_row}")
    print(f"Columns: {sheet.max_column}")
    print("========================================")

    # --------------------------------------------------------
    # Validate columns
    # --------------------------------------------------------

    if sheet.max_column < 4:
        raise ValueError(
            "Excel file must contain at least 4 columns: A, B, C, D."
        )

    # --------------------------------------------------------
    # Row 1 = Header
    # --------------------------------------------------------

    first_data_row = 2
    last_data_row = sheet.max_row

    if last_data_row < first_data_row:
        print("No data rows found.")
        workbook.save(output_file)
        return

    # --------------------------------------------------------
    # Read all complete rows
    # --------------------------------------------------------

    rows = []

    for row_number in range(
        first_data_row,
        last_data_row + 1
    ):

        row_data = read_row(
            sheet,
            row_number
        )

        rows.append(row_data)

    # --------------------------------------------------------
    # Ranking
    #
    # B = Gold
    # C = Silver
    # D = Bronze
    #
    # Lower number = better rank
    #
    # Priority:
    #
    # 1. Gold
    # 2. Silver
    # 3. Bronze
    # --------------------------------------------------------

    rows.sort(
        key=lambda row: (
            numeric_value(row["data"][1]["value"]),  # B - Gold
            numeric_value(row["data"][2]["value"]),  # C - Silver
            numeric_value(row["data"][3]["value"]),  # D - Bronze
        )
    )

    # --------------------------------------------------------
    # Write sorted rows back
    # --------------------------------------------------------

    for new_row_number, row_data in enumerate(
        rows,
        start=first_data_row
    ):

        write_row(
            sheet,
            new_row_number,
            row_data
        )

    # --------------------------------------------------------
    # Save result
    # --------------------------------------------------------

    workbook.save(output_file)

    print()
    print("Sorting completed successfully.")
    print(f"Saved to: {output_file}")
    print()
    print("Ranking rules:")
    print("1. Gold   (B) - ascending")
    print("2. Silver (C) - ascending")
    print("3. Bronze (D) - ascending")
    print()
    print("All columns A-D were moved together.")
    print("========================================")


# ============================================================
# Main
# ============================================================

if __name__ == "__main__":
    sort_excel(
        INPUT_FILE,
        OUTPUT_FILE
    )

